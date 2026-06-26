#!/usr/bin/env python3
"""build_master.py — 生成单一汇总工作簿(航线真实数据 + 运价模拟数据)。
流程: 1) node crawl.mjs / crawl_transit.mjs  2) python3 gen_mock_rates.py  3) python3 build_master.py
拿到真实运价后,替换 ocean_edges_MOCK.csv 数值,重跑本脚本即可。"""
import json, csv, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
routes=json.load(open('cosco_routes_raw.json'))['routes']; transit=json.load(open('cosco_transit_raw.json'))
TRADE={11:('跨太平洋航线',),12:('欧洲航线',),13:('欧地支线',),14:('亚太航线',),15:('拉非航线',),16:('东南亚及南亚航线',),17:('大西洋航线',),18:('中美洲航线',)}
DIR={'E':'东行/出口','W':'西行/进口','S':'南行/出口','N':'北行/进口'}
clean=lambda h: re.sub(r'<[^>]+>','',re.sub(r'<br\s*/?>','\n',h or '',flags=re.I)).strip()
def legs(x):
    c=x['callPort'].get('content') if x['callPort'] else None
    if not c: return []
    if isinstance(c,dict): c=[c]
    return [{**p,'_d':b.get('direction')} for b in c for p in b.get('ports',[])]
wb=Workbook(); HDR=PatternFill('solid',fgColor='2F75B5'); HF=Font(color='FFFFFF',bold=True)
def head(ws,n):
    for c in range(1,n+1): ws.cell(1,c).fill=HDR; ws.cell(1,c).font=HF
    ws.freeze_panes='A2'; ws.auto_filter.ref=f"A1:{get_column_letter(n)}1"
ws0=wb.active; ws0.title='说明 README'
for r in [['COSCO 航线 + 订舱运价 汇总'],['抓取日期','2026-06-23'],[''],
 ['① 港口挂靠明细','真实-航线路线表'],['② 运输时间表','真实-起运港→目的港天数'],['③ 航线索引','真实-193航线+优势'],
 ['④ 运价表 Rates','模拟MOCK-锚定 上海→Chicago $175/运输天'],[''],
 ['⚠️运价为模拟','替换 ocean_edges_MOCK.csv 后重跑;逻辑不变'],
 ['quote_optimizer.py','输入A→B,按价格/天数排序输出方案(含中转/箱型/门到门)']]: ws0.append(r)
ws0.column_dimensions['A'].width=30; ws0.column_dimensions['B'].width=60; ws0['A1'].font=Font(bold=True,size=14,color='2F75B5')
ws1=wb.create_sheet('① 港口挂靠明细 Port Calls')
ws1.append(['航线大类','航线组','航线代码','航线名称','方向','顺序','港口','码头','ETA','ETA累计天','ETD','ETD累计天','进出口'])
for x in sorted(routes,key=lambda r:(r['tradeUuid'],r['groupCn'] or '',r['serLpCode'] or '')):
    tc=TRADE.get(x['tradeUuid'],('',))[0]
    for p in legs(x): ws1.append([tc,x['groupCn'],x['serLpCode'],x['nameCn'],DIR.get(p.get('_d'),''),p.get('portOrder'),p.get('callPort'),p.get('callPortTerminal'),p.get('callPortEta'),p.get('callPortEtaTime'),p.get('callPortEtd'),p.get('callPortEtdTime'),p.get('isImportExport')])
head(ws1,13)
ws2=wb.create_sheet('② 运输时间表 Transit Times'); ws2.append(['航线大类','航线代码','航线名称','方向','起运港 POL','目的港 POD','运输天数']); D2={'E':'出口','W':'进口','S':'出口','N':'进口'}
for x in sorted(routes,key=lambda r:(r['tradeUuid'],r['serLpCode'] or '')):
    tc=TRADE.get(x['tradeUuid'],('',))[0]; t=transit.get(x['serLpCode']) or {}
    for side in ('loopExport','loopImport'):
        for blk in (t.get(side) or []):
            for pol in blk.get('Pol',[]):
                for pod in pol.get('pod',[]): ws2.append([tc,x['serLpCode'],x['nameCn'],D2.get(blk.get('direction'),''),pol.get('polName'),pod.get('podName'),int(pod['time']) if str(pod.get('time') or '').isdigit() else pod.get('time')])
head(ws2,7)
ws3=wb.create_sheet('③ 航线索引 Route Index'); ws3.append(['航线大类','航线组','航线代码','航线名称','挂靠港数','起始港','终点港','全程天','航线优势'])
for x in sorted(routes,key=lambda r:(r['tradeUuid'],r['groupCn'] or '',r['serLpCode'] or '')):
    tc=TRADE.get(x['tradeUuid'],('',))[0]; lg=legs(x); ports=[p['callPort'] for p in lg]; times=[int(p['callPortEtaTime']) for p in lg if str(p.get('callPortEtaTime') or '').isdigit()]
    ws3.append([tc,x['groupCn'],x['serLpCode'],x['nameCn'],len(lg),ports[0] if ports else '',ports[-1] if ports else '',max(times) if times else '',clean(x['advantage'].get('advantageContext') if x['advantage'] else '')])
head(ws3,9); ws3.column_dimensions['I'].width=55
ws4=wb.create_sheet('④ 运价表 Rates (MOCK)')
ws4.append(['集装箱类别','箱型','起运港','目的地','海运费USD','附加费USD','起运拖车USD','目的拖车USD',
            '港到港PP_USD','门到港DP_USD','港到门PD_USD','门到门DD_USD','运输天数','运价有效期','数据来源'])
door={r['港口']:r for r in csv.DictReader(open('door_charges_MOCK.csv'))}; CAT={'20GP':'干货箱','40GP':'干货箱','40HQ':'干货箱','40RF':'冷藏箱','40OT':'开顶箱','40FR':'框架箱'}
for r in csv.DictReader(open('ocean_edges_MOCK.csv')):
    box=r['箱型']; sz='20' if box.startswith('20') else '40'; A,B=r['起运港'],r['目的港']; oc=int(r['海运费USD']); su=int(r['附加费USD'])
    ot=int(door.get(A,{}).get(f'起运拖车{sz}USD',0) or 0); dt=int(door.get(B,{}).get(f'目的拖车{sz}USD',0) or 0)
    pp=oc+su  # 港到港 = 海运费+附加费;门到门/门到港/港到门 在此基础上加拖车
    ws4.append([CAT[box],box,A,B,oc,su,ot,dt, pp, pp+ot, pp+dt, pp+ot+dt, int(r['运输天数']),r['运价有效期'],'MOCK'])
head(ws4,15)
wb.save('COSCO_航线与运价_汇总_2026-06-23.xlsx'); print('saved unified workbook,',len(wb.sheetnames),'tabs')
