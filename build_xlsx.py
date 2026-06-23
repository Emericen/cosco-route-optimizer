import json, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

d=json.load(open('cosco_routes_raw.json'))
routes=d['routes']
transit=json.load(open('cosco_transit_raw.json'))

TRADE={11:('跨太平洋航线','Trans-Pacific'),12:('欧洲航线','Europe'),13:('欧地支线','Europe Feeder'),
14:('亚太航线','Asian Pacific'),15:('拉非航线','Latin America/Africa'),16:('东南亚及南亚航线','Southeast & South Asia'),
17:('大西洋航线','Trans-Atlantic'),18:('中美洲航线','Latin America Regional Service')}
DIR={'E':'东行/出口 Export','W':'西行/进口 Import','S':'南行/出口 Export','N':'北行/进口 Import'}

def clean(html):
    if not html: return ''
    t=re.sub(r'<br\s*/?>','\n',html,flags=re.I)
    t=re.sub(r'<[^>]+>','',t)
    return t.strip()

def legs(x):
    cp=x['callPort']; c=cp.get('content') if cp else None
    if not c: return []
    if isinstance(c,dict): c=[c]
    out=[]
    for b in c: out+=b.get('ports',[])
    return out

wb=Workbook()
HDR=PatternFill('solid',fgColor='2F75B5'); HF=Font(color='FFFFFF',bold=True,size=11)
THIN=Side(style='thin',color='D9D9D9'); BORD=Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
WRAP=Alignment(wrap_text=True,vertical='top'); CEN=Alignment(horizontal='center',vertical='top')

def style_header(ws,ncol):
    for c in range(1,ncol+1):
        cell=ws.cell(1,c); cell.fill=HDR; cell.font=HF; cell.alignment=Alignment(horizontal='center',vertical='center')
    ws.freeze_panes='A2'; ws.auto_filter.ref=f"A1:{get_column_letter(ncol)}1"

# ---- Sheet 1: Port Calls (long/tidy - the model-ready table) ----
ws=wb.active; ws.title='港口挂靠明细 Port Calls'
cols=['航线大类','Trade','航线组 Group','航线代码 Code','航线名称 Route','方向 Direction',
      '挂靠顺序 Seq','港口 Port','码头 Terminal','ETA(星期)','ETA累计天','ETD(星期)','ETD累计天','进出口 In/Out','备注']
ws.append(cols)
for x in sorted(routes,key=lambda r:(r['tradeUuid'],r['groupCn'] or '',r['serLpCode'] or '')):
    tc,te=TRADE.get(x['tradeUuid'],(str(x['tradeUuid']),''))
    for p in legs(x):
        ws.append([tc,te,x['groupCn'],x['serLpCode'],x['nameCn'],DIR.get(p.get('direction'),p.get('direction') or ''),
            p.get('portOrder'),p.get('callPort'),p.get('callPortTerminal'),p.get('callPortEta'),p.get('callPortEtaTime'),
            p.get('callPortEtd'),p.get('callPortEtdTime'),p.get('isImportExport'),p.get('callPortRemark') or ''])
widths=[16,22,22,12,14,18,9,16,30,10,9,10,9,9,18]
for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
style_header(ws,len(cols))

# ---- Sheet 2: Route Index (one row per route + advantages) ----
ws2=wb.create_sheet('航线索引 Route Index')
cols2=['航线大类','Trade','航线组 Group','航线代码 Code','航线名称 Route','英文名 Name(EN)',
       '挂靠港数','起始港 Origin','终点港 Final','全程(天)','航线优势 Advantages']
ws2.append(cols2)
for x in sorted(routes,key=lambda r:(r['tradeUuid'],r['groupCn'] or '',r['serLpCode'] or '')):
    tc,te=TRADE.get(x['tradeUuid'],(str(x['tradeUuid']),''))
    lg=legs(x)
    ports=[p['callPort'] for p in lg]
    times=[int(p['callPortEtaTime']) for p in lg if str(p.get('callPortEtaTime') or '').isdigit()]
    adv=clean(x['advantage'].get('advantageContext') if x['advantage'] else '')
    ws2.append([tc,te,x['groupCn'],x['serLpCode'],x['nameCn'],x['nameEn'],len(lg),
        ports[0] if ports else '', ports[-1] if ports else '', max(times) if times else '', adv])
w2=[16,22,22,12,14,16,9,14,14,9,60]
for i,w in enumerate(w2,1): ws2.column_dimensions[get_column_letter(i)].width=w
for row in ws2.iter_rows(min_row=2,min_col=11,max_col=11): row[0].alignment=WRAP
style_header(ws2,len(cols2))

# ---- Sheet 3: Transit Times (运输时间表 - POL->POD transit days) ----
ws4=wb.create_sheet('运输时间表 Transit Times')
cols4=['航线大类','航线组 Group','航线代码 Code','航线名称 Route','方向 Direction',
       '起运港 POL','目的港 POD','运输天数 Transit Days']
ws4.append(cols4)
DIR2={'E':'出口 Export','W':'进口 Import','S':'出口 Export','N':'进口 Import'}
rinfo={x['serLpCode']:x for x in routes}
tpairs=0
for x in sorted(routes,key=lambda r:(r['tradeUuid'],r['groupCn'] or '',r['serLpCode'] or '')):
    tc,te=TRADE.get(x['tradeUuid'],(str(x['tradeUuid']),''))
    t=transit.get(x['serLpCode']) or {}
    for side in ('loopExport','loopImport'):
        blocks=t.get(side) or []
        for blk in blocks:
            dlabel=DIR2.get(blk.get('direction'),blk.get('direction') or '')
            for pol in blk.get('Pol',[]):
                for pod in pol.get('pod',[]):
                    ws4.append([tc,x['groupCn'],x['serLpCode'],x['nameCn'],dlabel,
                                pol.get('polName'),pod.get('podName'),
                                int(pod['time']) if str(pod.get('time') or '').lstrip('-').isdigit() else pod.get('time')])
                    tpairs+=1
w4=[16,22,12,14,14,20,20,14]
for i,w in enumerate(w4,1): ws4.column_dimensions[get_column_letter(i)].width=w
style_header(ws4,len(cols4))

# ---- Sheet 4: README ----
ws3=wb.create_sheet('说明 README',0)
notes=[
 ('COSCO 全航线汇总 / COSCO Shipping – All Routes',''),
 ('数据来源','world.lines.coscoshipping.com (官方接口 /homeapiak/routeService)'),
 ('抓取日期','2026-06-23'),
 ('航线大类 (Trades)','8 个'),('航线组 (Groups)','35 个'),
 ('唯一航线 (Unique routes)','193 条'),
 ('方向序列 (出口+进口 directional sequences)','385 条 (网站“419”约指此口径，含双向)'),
 ('港口挂靠记录 (Port-call rows)','1865 条'),
 ('港口对运输时间 (POL→POD transit pairs)','约 4984 条'),
 ('',''),
 ('工作表 Sheets',''),
 ('① 港口挂靠明细 Port Calls','【航线路线表】每行=一个挂靠港口。含港口、码头、ETA/ETD、累计天数、方向、进出口。'),
 ('② 运输时间表 Transit Times','【运输时间表】每行=一个 起运港→目的港 组合及其运输天数。可直接作为 A→B 路线优化的边权。'),
 ('③ 航线索引 Route Index','每行=一条航线。含起止港、全程天数、航线优势。'),
 ('',''),
 ('字段说明 Notes',''),
 ('ETA累计天/ETD累计天','自该航线起始港(第0天)算起的累计天数。'),
 ('运输天数 Transit Days','某起运港(POL)到某目的港(POD)的运输时间(天) → A→B 优化模型的核心输入。'),
 ('方向 Direction','东行/南行=出口(Export);西行/北行=进口(Import)。同一航线两个方向分别列出。'),
 ('挂靠顺序 Seq','港口在该航线该方向上的先后顺序(从0开始)。'),
]
for r_ in notes: ws3.append(list(r_))
ws3.column_dimensions['A'].width=42; ws3.column_dimensions['B'].width=80
ws3['A1'].font=Font(bold=True,size=14,color='2F75B5')
for r in range(2,len(notes)+1):
    ws3.cell(r,1).font=Font(bold=True); ws3.cell(r,2).alignment=WRAP

out="COSCO_全航线汇总_2026-06-23.xlsx"
wb.save(out)
print('saved:',out)
print('Port Calls rows:',ws.max_row-1,'| Transit rows:',ws4.max_row-1,'| Route Index rows:',ws2.max_row-1)
